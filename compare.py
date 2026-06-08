"""
Compares 2026_master.csv against 2025 Faculty List With SCOPUS IDs.xlsx.
Produces output/2026_master.xlsx with two sheets:
  - "2026 Master"  : all 2026 faculty data
  - "comparison"   : faculty present in one year but not the other

Matching key: normalized full name + normalized university base name.
"""

import csv
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

BASE_DIR   = Path(__file__).parent
EXCEL_2025 = BASE_DIR / "2025 Faculty List With SCOPUS IDs.xlsx"
CSV_2026   = BASE_DIR / "output" / "2026_master.csv"
OUT_XLSX   = BASE_DIR / "output" / "2026_master.xlsx"

COMPARISON_COLS = [
    "status", "scopus_id", "name", "first_name", "last_name",
    "original_title", "department", "area", "university", "rank",
]

COMBINE_COLS = [
    "scopus_id", "suggest_id", "name", "original_title", "university", "area",
    "first_name", "last_name", "rank_2025", "rank_2026", "status",
]

COMBINE_HEADERS = [
    "ScopusAuthorID", "SuggestID", "Full Name", "Original Title", "University", "Original Area",
    "First Name", "Last Name", "2025 Rank", "2026 Rank", "Status",
]

TT_RANKS_2026 = {"professor", "associate", "assistant"}

# 2025 Excel uses "at City" suffixes that the 2026 config omits
UNIV_ALIASES: dict[str, str] = {
    "indiana university at bloomington":              "indiana university",
    "pennsylvania state university at state college": "pennsylvania state university",
    "university of michigan at ann arbor":            "university of michigan",
    "university of minnesota at twin cities":         "university of minnesota",
    "university of south carolina at columbia":       "university of south carolina",
    "university of washington at seattle":            "university of washington",
}

RANK_ORDER = {"assistant": 1, "associate": 2, "professor": 3}


def norm_name(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").lower().strip())


def norm_univ(s: str) -> str:
    """Strip school/college suffix in parentheses, lowercase, strip, apply aliases."""
    base = re.sub(r"\s*\(.*", "", s or "")
    result = base.lower().strip()
    return UNIV_ALIASES.get(result, result)


def norm_area(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").lower().strip())


def fl_key(first: str, last: str) -> str:
    """Fallback match key: first word of first_name + last_name, lowercase.
    Handles middle-initial mismatches: 'James A. Best' vs 'James Best'."""
    fw = (first or "").split()[0].lower() if first else ""
    return fw + "|" + (last or "").lower().strip()


def load_2025() -> dict[tuple, dict]:
    wb = openpyxl.load_workbook(EXCEL_2025, read_only=True, data_only=True)
    ws = wb["2025 Master Copy"]
    faculty = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        raw_sid = str(row[3]).strip() if row[3] is not None else ""
        try:
            scopus_id = str(int(float(raw_sid))) if raw_sid and raw_sid not in ("None", "N/A", "") else ""
        except (ValueError, OverflowError):
            scopus_id = raw_sid
        full_name    = str(row[4] or "").strip()
        original_title = str(row[5] or "").strip()
        university   = str(row[6] or "").strip()
        department   = str(row[7] or "").strip()  # OriginalArea used as dept
        first_name   = str(row[8] or "").strip()
        last_name    = str(row[9] or "").strip()
        rank         = str(row[10] or "").strip()
        if not full_name:
            continue
        key = (norm_name(full_name), norm_univ(university))
        faculty[key] = {
            "scopus_id":     scopus_id,
            "name":          full_name,
            "first_name":    first_name,
            "last_name":     last_name,
            "original_title": original_title,
            "department":    department,
            "area":          department,  # 2025 has no separate area col
            "university":    university,
            "rank":          rank,
        }
    wb.close()
    return faculty


def load_2026() -> tuple[list[dict], dict[tuple, dict]]:
    rows = []
    index = {}
    with open(CSV_2026, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            rows.append(row)
            key = (norm_name(row["name"]), norm_univ(row["university"]))
            index[key] = row
    return rows, index


def style_header(ws, row_num: int, ncols: int, fill_hex: str):
    fill = PatternFill("solid", fgColor=fill_hex)
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def write_xlsx(rows_2026: list[dict], comparison_rows: list[dict]):
    wb = openpyxl.Workbook()

    # ── Sheet 1: 2026 Master ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "2026 Master"
    master_cols = ["scopus_id", "name", "first_name", "last_name",
                   "original_title", "department", "area", "university",
                   "email", "rank"]
    ws1.append(master_cols)
    style_header(ws1, 1, len(master_cols), "2E75B6")
    for r in rows_2026:
        ws1.append([r.get(c, "") for c in master_cols])

    # ── Sheet 2: comparison ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("comparison")
    ws2.append(COMPARISON_COLS)
    style_header(ws2, 1, len(COMPARISON_COLS), "70AD47")

    green_fill = PatternFill("solid", fgColor="E2EFDA")  # new in 2026
    red_fill   = PatternFill("solid", fgColor="FCE4D6")  # left / not in 2026

    for r in comparison_rows:
        ws2.append([r.get(c, "") for c in COMPARISON_COLS])
        row_idx = ws2.max_row
        fill = green_fill if r["status"] == "New in 2026" else red_fill
        for col in range(1, len(COMPARISON_COLS) + 1):
            ws2.cell(row=row_idx, column=col).fill = fill

    # Auto-width (approximate) for both sheets
    for ws in [ws1, ws2]:
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 50)

    wb.save(OUT_XLSX)
    print(f"Saved → {OUT_XLSX}")


def build_2026_name_index(rows_2026_all: list[dict]) -> tuple[dict, dict]:
    """Returns (name_index, fl_name_index).
    name_index   : { norm_name → [(disp_univ, norm_univ, rank)] }
    fl_name_index: { fl_key   → [(disp_univ, norm_univ, rank)] }  (fallback for middle-initial mismatches)
    """
    index:    dict[str, list] = {}
    fl_index: dict[str, list] = {}
    for row in rows_2026_all:
        entry = (row["university"], norm_univ(row["university"]), row.get("rank", ""))
        index.setdefault(norm_name(row["name"]), []).append(entry)
        flk = fl_key(row.get("first_name", ""), row.get("last_name", ""))
        fl_index.setdefault(flk, []).append(entry)
    return index, fl_index


def get_status(r25_rank: str, r26_rank: str,
               norm_n: str, norm_u: str,
               name_index: dict,
               fl_name_index: dict | None = None,
               r25_first: str = "", r25_last: str = "") -> str:
    r25 = r25_rank.strip().lower()
    r26 = r26_rank.strip().lower()

    if r25 and r26:
        r25_ord = RANK_ORDER.get(r25, 0)
        r26_ord = RANK_ORDER.get(r26, 0)
        if r26_ord > r25_ord:
            return "Promoted"
        if r26_ord < r25_ord:
            return "Demoted"
        return "Active"

    if not r25 and r26:
        return "New hire"

    # r25 filled, r26 blank — person left TT at this university
    matches = name_index.get(norm_n, [])
    # Fallback: try fl_key if exact name not found (handles middle-initial mismatches)
    if not matches and fl_name_index and r25_first:
        matches = fl_name_index.get(fl_key(r25_first, r25_last), [])

    # 1. Emeritus anywhere
    for disp_u, _, rank in matches:
        if str(rank).lower() == "emeritus":
            return "Retired (Emeritus)"

    # 2. TT at a different university — only if unambiguous (exactly one other school)
    tt_elsewhere = list({(disp_u, nu) for disp_u, nu, rank in matches
                         if nu != norm_u and str(rank).lower() in TT_RANKS_2026})
    if len(tt_elsewhere) == 1:
        return f"Moved to {tt_elsewhere[0][0]}"

    # 3. Non-TT at same university — use actual 2026 rank (e.g. Dean, Lecturer, etc.)
    same_univ_ranks = list(dict.fromkeys(
        rank for _, nu, rank in matches if nu == norm_u
    ))
    if same_univ_ranks:
        return same_univ_ranks[0]

    # 4. Non-TT at exactly one different university
    non_tt_elsewhere = list({(disp_u, nu) for disp_u, nu, _ in matches if nu != norm_u})
    if len(non_tt_elsewhere) == 1:
        return f"Moved to {non_tt_elsewhere[0][0]} (non-TT)"

    return "Unknown / Dismissed"


def build_combine(faculty_2025: dict, rows_2026_all: list[dict]) -> list[dict]:
    rows_2026_tt = [r for r in rows_2026_all if r.get("rank", "").lower() in TT_RANKS_2026]

    index_2026_tt: dict[tuple, dict] = {}
    # Secondary index keyed by (first_word_of_first_name|last_name, norm_univ)
    # Used when full-name match fails due to middle-initial formatting differences
    fl_index_2026_tt: dict[tuple, dict] = {}
    for row in rows_2026_tt:
        key = (norm_name(row["name"]), norm_univ(row["university"]))
        index_2026_tt[key] = row
        fl_k = (fl_key(row.get("first_name", ""), row.get("last_name", "")),
                norm_univ(row["university"]))
        fl_index_2026_tt.setdefault(fl_k, row)  # keep first; avoids overwriting on collision

    name_index, fl_name_index = build_2026_name_index(rows_2026_all)

    combine_rows = []
    matched_2026_keys: set[tuple] = set()

    for key, r25 in faculty_2025.items():
        r26 = index_2026_tt.get(key)
        norm_n, norm_u = key

        # Fallback: try first-word-of-first + last + univ (handles middle-initial diffs)
        if r26 is None:
            fl_k = (fl_key(r25.get("first_name", ""), r25.get("last_name", "")), norm_u)
            r26 = fl_index_2026_tt.get(fl_k)

        r26_rank = r26["rank"] if r26 else ""
        status = get_status(r25["rank"], r26_rank, norm_n, norm_u, name_index,
                            fl_name_index, r25.get("first_name", ""), r25.get("last_name", ""))

        # For "Moved to" rows, fill 2026 rank only if destination rank is TT
        if status.startswith("Moved to"):
            matches_for_rank = name_index.get(norm_n, []) or fl_name_index.get(
                fl_key(r25.get("first_name",""), r25.get("last_name","")), [])
            for _, nu, rank in matches_for_rank:
                if nu != norm_u and str(rank).lower() in TT_RANKS_2026:
                    r26_rank = rank
                    break

        combine_rows.append({
            "scopus_id":      r25.get("scopus_id", ""),   # strictly from 2025 xlsx
            "suggest_id":     "",                          # blank for 2025 people
            "name":           r25["name"],
            "original_title": r25["original_title"],
            "university":     r25["university"],
            "area":           r25["area"],
            "first_name":     r25["first_name"],
            "last_name":      r25["last_name"],
            "rank_2025":      r25["rank"],
            "rank_2026":      r26_rank,
            "status":         status,
        })

        if r26:
            matched_2026_keys.add(key)

    for key, r26 in index_2026_tt.items():
        if key not in matched_2026_keys:
            norm_n, norm_u = key
            status = get_status("", r26["rank"], norm_n, norm_u, name_index,
                                fl_name_index, r26.get("first_name", ""), r26.get("last_name", ""))
            combine_rows.append({
                "scopus_id":      "",                        # blank — not in 2025
                "suggest_id":     r26.get("scopus_id", ""),  # from 2026 pipeline
                "name":           r26["name"],
                "original_title": r26.get("original_title", ""),
                "university":     r26["university"],
                "area":           r26.get("area", ""),
                "first_name":     r26.get("first_name", ""),
                "last_name":      r26.get("last_name", ""),
                "rank_2025":      "",
                "rank_2026":      r26["rank"],
                "status":         status,
            })

    combine_rows.sort(key=lambda r: (r["university"], r["last_name"]))
    return combine_rows


def write_combine_sheet(combine_rows: list[dict]):
    """Open the existing 2026_master.xlsx and add/replace only the 'combine' sheet."""
    wb = openpyxl.load_workbook(OUT_XLSX)
    if "combine" in wb.sheetnames:
        del wb["combine"]

    ws = wb.create_sheet("combine")
    ws.append(COMBINE_HEADERS)
    style_header(ws, 1, len(COMBINE_HEADERS), "7030A0")

    for r in combine_rows:
        ws.append([r.get(c, "") for c in COMBINE_COLS])

    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 50)

    wb.save(OUT_XLSX)
    print(f"Saved → {OUT_XLSX}")


VERIFY_COLS    = COMBINE_COLS + ["verify_reason"]
VERIFY_HEADERS = COMBINE_HEADERS + ["Verify Reason"]


def build_verify_rows(combine_rows: list[dict], faculty_2025: dict) -> list[dict]:
    """Return rows whose ScopusAuthorID needs manual verification."""
    verify = []
    for r in combine_rows:
        sid_scopus  = str(r.get("scopus_id")  or "").strip()
        sid_suggest = str(r.get("suggest_id") or "").strip()
        is_new_hire = not r.get("rank_2025", "").strip()

        if is_new_hire:
            if not sid_suggest:
                reason = "No Scopus ID found"
            else:
                reason = "ID from 2026 lookup — not in 2025 data"
        else:
            if not sid_scopus:
                reason = "No Scopus ID found"
            else:
                continue   # ID came from 2025 xlsx — confirmed

        verify.append({**r, "verify_reason": reason})

    verify.sort(key=lambda r: (r["university"], r["last_name"]))
    return verify


def write_verify_sheet(verify_rows: list[dict]):
    """Open existing xlsx and add/replace the verify_ID sheet."""
    wb = openpyxl.load_workbook(OUT_XLSX)
    if "verify_ID" in wb.sheetnames:
        del wb["verify_ID"]

    ws = wb.create_sheet("verify_ID")
    ws.append(VERIFY_HEADERS)
    style_header(ws, 1, len(VERIFY_HEADERS), "FFC000")   # amber

    yellow_fill = PatternFill("solid", fgColor="FFFF99")

    for r in verify_rows:
        ws.append([r.get(c, "") for c in VERIFY_COLS])
        row_idx = ws.max_row
        for col in range(1, len(VERIFY_COLS) + 1):
            ws.cell(row=row_idx, column=col).fill = yellow_fill

    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 50)

    wb.save(OUT_XLSX)
    print(f"Saved → {OUT_XLSX}")


def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--rebuild", action="store_true",
        help="Fully regenerate the xlsx from scratch (overwrites all sheets)"
    )
    args = parser.parse_args()

    print("Loading 2025 data...")
    faculty_2025 = load_2025()
    print(f"  {len(faculty_2025)} faculty in 2025")

    print("Loading 2026 data...")
    rows_2026, faculty_2026 = load_2026()
    print(f"  {len(rows_2026)} faculty in 2026")

    if args.rebuild:
        new_in_2026 = [{"status": "New in 2026", **row} for key, row in faculty_2026.items() if key not in faculty_2025]
        not_in_2026 = [{"status": "Not in 2026", **row} for key, row in faculty_2025.items() if key not in faculty_2026]
        new_in_2026.sort(key=lambda r: (r["university"], r["last_name"]))
        not_in_2026.sort(key=lambda r: (r["university"], r["last_name"]))
        comparison_rows = new_in_2026 + not_in_2026
        print(f"\nNew in 2026 (not in 2025): {len(new_in_2026)}")
        print(f"Not in 2026 (was in 2025): {len(not_in_2026)}")
        print(f"Total comparison rows:     {len(comparison_rows)}")
        print("\nWriting full Excel file...")
        write_xlsx(rows_2026, comparison_rows)

    print("\nBuilding combine sheet...")
    combine_rows = build_combine(faculty_2025, rows_2026)
    print(f"  {len(combine_rows)} rows in combine")
    write_combine_sheet(combine_rows)

    print("\nBuilding verify_ID sheet...")
    verify_rows = build_verify_rows(combine_rows, faculty_2025)
    print(f"  {len(verify_rows)} rows to verify")
    write_verify_sheet(verify_rows)


if __name__ == "__main__":
    main()
