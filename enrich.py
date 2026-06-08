"""
Looks up ScopusAuthorIDs for faculty rows that have no scopus_id yet.
Calls the Scopus Author Search API with first name + last name + affiliation.

Results:
  - 1 match found   → auto-filled
  - Multiple matches → printed for manual review
  - 0 matches        → flagged as not found

Usage (standalone):
  python enrich.py                        # process all CSVs in output/
  python enrich.py --index 01            # one university
  python enrich.py --verify              # scan verify_ID sheet in 2026_master.xlsx
"""

import argparse
import csv
import json
import os
import re
import time
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import requests

CONFIG_PATH = Path(__file__).parent / "config" / "universities.json"
OUTPUT_DIR  = Path(__file__).parent / "output"
SCOPUS_URL  = "https://api.elsevier.com/content/search/author"
VERIFY_XLSX = OUTPUT_DIR / "2026_master.xlsx"

# Column order for verify_ID sheet — must match compare.py VERIFY_COLS / VERIFY_HEADERS
_VERIFY_COLS = [
    "scopus_id", "name", "original_title", "university", "area",
    "first_name", "last_name", "rank_2025", "rank_2026", "status", "verify_reason",
]
_VERIFY_HDRS = [
    "ScopusAuthorID", "Full Name", "Original Title", "University", "Original Area",
    "First Name", "Last Name", "2025 Rank", "2026 Rank", "Status", "Verify Reason",
]
_VERIFY_HDR_TO_COL = dict(zip(_VERIFY_HDRS, _VERIFY_COLS))

# Broad Scopus subject-area codes keyed by area keyword
_AREA_SUBJ: dict[str, str] = {
    "accounting":  "BUSI",
    "finance":     "BUSI",
    "marketing":   "BUSI",
    "management":  "BUSI",
    "economics":   "ECON",
    "operations":  "DECI",
    "technology":  "DECI",
    "information": "DECI",
}


class QuotaExhaustedError(Exception):
    pass


# ---------------------------------------------------------------------------
# Load API key from environment (set in .venv/bin/activate)
# ---------------------------------------------------------------------------
def load_api_key() -> str:
    key = os.environ.get("SCOPUS_API_KEY", "")
    if not key:
        raise RuntimeError(
            "SCOPUS_API_KEY not found. Make sure your .venv is activated: "
            "source .venv/bin/activate"
        )
    return key


# ---------------------------------------------------------------------------
# Scopus Author Search
# ---------------------------------------------------------------------------
def search_author(first: str, last: str, affiliation: str, api_key: str) -> list[dict]:
    query = f'AUTHFIRST("{first}") AND AUTHLAST("{last}") AND AFFIL("{affiliation}")'
    params = {"query": query, "count": 5, "field": "identifier,preferred-name,affiliation-current"}
    headers = {"X-ELS-APIKey": api_key, "Accept": "application/json"}

    try:
        resp = requests.get(SCOPUS_URL, params=params, headers=headers, timeout=15)
        if resp.status_code == 429:
            raise QuotaExhaustedError()
        resp.raise_for_status()
        data = resp.json()
        entries = data.get("search-results", {}).get("entry", [])
        if entries and "error" in entries[0]:
            return []
        return entries
    except requests.RequestException as e:
        print(f"    API error for {first} {last}: {e}")
        return []


def parse_author_id(entry: dict) -> str:
    raw = entry.get("dc:identifier", "")
    return raw.replace("AUTHOR_ID:", "").strip()


def parse_affil(entry: dict) -> str:
    affil = entry.get("affiliation-current", {})
    if isinstance(affil, list):
        affil = affil[0] if affil else {}
    return affil.get("affiliation-name", "")


def parse_display_name(entry: dict) -> str:
    pn = entry.get("preferred-name", {})
    return f"{pn.get('given-name', '')} {pn.get('surname', '')}".strip()


# ---------------------------------------------------------------------------
# Process one CSV (normal pipeline mode)
# ---------------------------------------------------------------------------
def enrich(csv_path: Path, affiliation: str, api_key: str):
    with open(csv_path, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    fieldnames = list(rows[0].keys()) if rows else []
    missing = [r for r in rows if not r.get("scopus_id")]

    if not missing:
        print("  All rows already have Scopus IDs — nothing to do.")
        return

    print(f"  Looking up {len(missing)} faculty without Scopus ID...")

    auto_filled = 0
    multi_match = []
    not_found   = []

    for row in missing:
        first = row["first_name"]
        last  = row["last_name"]
        entries = search_author(first, last, affiliation, api_key)

        if len(entries) == 1:
            row["scopus_id"] = parse_author_id(entries[0])
            auto_filled += 1

        elif len(entries) > 1:
            candidates = [
                (parse_author_id(e), parse_display_name(e), parse_affil(e))
                for e in entries
            ]
            multi_match.append((row["name"], candidates))

        else:
            not_found.append(row["name"])

        time.sleep(1.0)  # stay within API rate limit

    # Write updated CSV (partial progress saved even if quota hits later)
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    print(f"  Auto-filled:    {auto_filled}")
    print(f"  Multiple hits:  {len(multi_match)} (pick manually below)")
    print(f"  Not found:      {len(not_found)}")

    if multi_match:
        print(f"\n  --- Multiple candidates — open the CSV and fill scopus_id manually ---")
        for name, candidates in multi_match:
            print(f"\n  Faculty: {name}")
            for sid, dname, aff in candidates:
                print(f"    ID={sid:>15s}  name={dname:<30s}  affil={aff}")

    if not_found:
        print(f"\n  --- Not found in Scopus (may have no publications yet) ---")
        for n in not_found:
            print(f"    {n}")


# ---------------------------------------------------------------------------
# Verify mode helpers
# ---------------------------------------------------------------------------
def _area_to_subj(area: str) -> str:
    """Map our area label → a Scopus SUBJAREA code, or '' if no mapping."""
    low = area.lower()
    for kw, code in _AREA_SUBJ.items():
        if kw in low:
            return code
    return ""


def _alt_firsts(first: str) -> list[str]:
    """
    Alternative AUTHFIRST terms for compound/initial names.
    "E. Geoffrey" → ["Geoffrey"]
    "Mary Ann"    → ["Mary"]
    "Jean-Pierre" → ["Jean Pierre"]
    """
    alts: list[str] = []
    parts = first.split()
    if len(parts) >= 2 and re.match(r"^[A-Z]\.$", parts[0]):
        alts.append(" ".join(parts[1:]))   # drop leading initial
    elif len(parts) >= 2:
        alts.append(parts[0])              # just first token
    if "-" in first:
        alts.append(first.replace("-", " "))
    return alts


def _raw_search(query: str, api_key: str) -> list[dict]:
    """Low-level search: send a raw Scopus query string, return entries or []."""
    params  = {"query": query, "count": 5,
                "field": "identifier,preferred-name,affiliation-current"}
    headers = {"X-ELS-APIKey": api_key, "Accept": "application/json"}
    try:
        resp = requests.get(SCOPUS_URL, params=params, headers=headers, timeout=15)
        if resp.status_code == 429:
            raise QuotaExhaustedError()
        resp.raise_for_status()
        data    = resp.json()
        entries = data.get("search-results", {}).get("entry", [])
        return [] if (entries and "error" in entries[0]) else entries
    except QuotaExhaustedError:
        raise
    except requests.RequestException as e:
        print(f"    API error: {e}")
        return []


def search_deep(first: str, last: str, affiliation: str, area: str,
                api_key: str) -> tuple[list[dict], str]:
    """
    Multi-pass author search returning (entries, strategy_label).

    Pass 1 — standard: AUTHFIRST + AUTHLAST + AFFIL
    Pass 1b — if multiple: add SUBJAREA to narrow
    Pass 2 — alternative first-name form if 0 results
    """
    # Pass 1
    entries = search_author(first, last, affiliation, api_key)

    if len(entries) == 1:
        return entries, "standard"

    if len(entries) > 1:
        subj = _area_to_subj(area)
        if subj:
            time.sleep(0.5)
            q2 = (f'AUTHFIRST("{first}") AND AUTHLAST("{last}") '
                  f'AND AFFIL("{affiliation}") AND SUBJAREA({subj})')
            narrowed = _raw_search(q2, api_key)
            if len(narrowed) == 1:
                return narrowed, f"standard+subj({subj})"
            if narrowed:
                return narrowed, f"multi({len(narrowed)})+subj({subj})"
        return entries, f"multi({len(entries)})"

    # Pass 2 — try alternate first-name forms (still with affiliation)
    for alt in _alt_firsts(first):
        time.sleep(0.5)
        e2 = search_author(alt, last, affiliation, api_key)
        if len(e2) == 1:
            return e2, f"alt_first={alt}"
        if len(e2) > 1:
            subj = _area_to_subj(area)
            if subj:
                time.sleep(0.5)
                q3 = (f'AUTHFIRST("{alt}") AND AUTHLAST("{last}") '
                      f'AND AFFIL("{affiliation}") AND SUBJAREA({subj})')
                e3 = _raw_search(q3, api_key)
                if len(e3) == 1:
                    return e3, f"alt_first={alt}+subj({subj})"
                if e3:
                    return e3, f"alt_first={alt}/multi({len(e3)})"
            return e2, f"alt_first={alt}/multi({len(e2)})"

    # Pass 3 — drop affiliation (handles faculty whose Scopus affil lags their move)
    subj = _area_to_subj(area)
    for search_first in [first] + _alt_firsts(first):
        time.sleep(0.5)
        q_noaffil = f'AUTHFIRST("{search_first}") AND AUTHLAST("{last}")'
        if subj:
            q_noaffil += f" AND SUBJAREA({subj})"
        e_na = _raw_search(q_noaffil, api_key)
        if len(e_na) == 1:
            return e_na, f"no_affil+subj({subj})" if subj else "no_affil"
        if len(e_na) > 1:
            return e_na, f"no_affil/multi({len(e_na)})"

    return [], "not_found"


# ---------------------------------------------------------------------------
# Read / write verify_ID sheet
# ---------------------------------------------------------------------------
def load_verify_rows(xlsx_path: Path) -> list[dict]:
    """Return rows from the verify_ID sheet as dicts with internal field names."""
    wb   = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws   = wb["verify_ID"]
    cols: list[str] = []
    rows: list[dict] = []
    for i, row_vals in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            cols = [_VERIFY_HDR_TO_COL.get(str(v or "").strip(),
                                            str(v or "").strip())
                    for v in row_vals]
            continue
        rows.append({cols[j]: (str(v) if v is not None else "")
                     for j, v in enumerate(row_vals)})
    wb.close()
    return rows


def _style_header_row(ws, ncols: int, fill_hex: str):
    fill = PatternFill("solid", fgColor=fill_hex)
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = fill
        cell.alignment = Alignment(horizontal="center")


def update_sheets(xlsx_path: Path, updated_rows: list[dict]):
    """
    Rewrite verify_ID sheet with updated_rows.
    For rows marked Resolved, also patch scopus_id in the combine sheet.
    """
    resolved: dict[tuple, str] = {
        (r["name"], r["university"]): r["scopus_id"]
        for r in updated_rows
        if r.get("verify_reason", "").startswith("Resolved:")
    }

    wb = openpyxl.load_workbook(xlsx_path)

    # ── verify_ID ────────────────────────────────────────────────────────
    if "verify_ID" in wb.sheetnames:
        del wb["verify_ID"]
    ws_v = wb.create_sheet("verify_ID")
    ws_v.append(_VERIFY_HDRS)
    _style_header_row(ws_v, len(_VERIFY_HDRS), "FFC000")

    yellow_fill = PatternFill("solid", fgColor="FFFF99")
    for r in updated_rows:
        ws_v.append([r.get(c, "") for c in _VERIFY_COLS])
        ri = ws_v.max_row
        for col in range(1, len(_VERIFY_COLS) + 1):
            ws_v.cell(ri, col).fill = yellow_fill

    for col_cells in ws_v.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws_v.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 50)

    # ── patch combine ────────────────────────────────────────────────────
    if resolved and "combine" in wb.sheetnames:
        ws_c  = wb["combine"]
        hdrs  = [ws_c.cell(1, c).value for c in range(1, ws_c.max_column + 1)]
        try:
            sid_col  = hdrs.index("ScopusAuthorID") + 1
            name_col = hdrs.index("Full Name")      + 1
            univ_col = hdrs.index("University")     + 1
        except ValueError:
            sid_col = name_col = univ_col = None

        if sid_col and name_col and univ_col:
            for ri in range(2, ws_c.max_row + 1):
                name = str(ws_c.cell(ri, name_col).value or "")
                univ = str(ws_c.cell(ri, univ_col).value or "")
                new_id = resolved.get((name, univ))
                if new_id:
                    ws_c.cell(ri, sid_col).value = new_id

    wb.save(xlsx_path)
    print(f"  Saved → {xlsx_path}")


# ---------------------------------------------------------------------------
# Verify entry point
# ---------------------------------------------------------------------------
def run_verify(xlsx_path: Path, api_key: str):
    """Scan every row in verify_ID, try to resolve Scopus IDs, write results back."""
    if not xlsx_path.exists():
        print(f"  File not found: {xlsx_path}")
        return

    rows = load_verify_rows(xlsx_path)
    print(f"  {len(rows)} rows in verify_ID")

    already_done = sum(1 for r in rows if r.get("verify_reason", "").startswith("Resolved:"))
    to_process   = [r for r in rows if not r.get("verify_reason", "").startswith("Resolved:")]
    print(f"  Already resolved: {already_done}")
    print(f"  To process:       {len(to_process)}")

    auto_filled  = 0
    multi_match: list[tuple] = []
    not_found_ct = 0

    for i, row in enumerate(to_process, 1):
        first = row.get("first_name", "")
        last  = row.get("last_name", "")
        univ  = row.get("university", "")
        area  = row.get("area", "")
        name  = row.get("name", "")

        # Strip parenthetical school name: "Arizona State University (W.P. Carey...)" → "Arizona State University"
        affil = re.sub(r"\s*\(.*", "", univ).strip()

        print(f"  [{i}/{len(to_process)}] {name} @ {univ}", end=" ... ", flush=True)

        try:
            entries, strategy = search_deep(first, last, affil, area, api_key)
        except QuotaExhaustedError:
            print(f"\n  QUOTA EXHAUSTED at [{i}] {name}.")
            print(f"  Re-run with: python enrich.py --verify  (progress is saved)")
            print(f"  Quota resets weekly (5,000 calls/week). Use a second API key or wait for the weekly reset.")
            break

        if len(entries) == 1:
            new_id = parse_author_id(entries[0])
            row["scopus_id"]      = new_id
            row["verify_reason"]  = f"Resolved: {strategy}"
            auto_filled += 1
            print(f"✓ {new_id}")
        elif len(entries) > 1:
            candidates = [
                (parse_author_id(e), parse_display_name(e), parse_affil(e))
                for e in entries
            ]
            multi_match.append((name, univ, candidates))
            print(f"multiple ({len(entries)})")
        else:
            not_found_ct += 1
            print("not found")

        time.sleep(1.0)

    print(f"\n  Writing updated sheets...")
    update_sheets(xlsx_path, rows)

    print(f"\n  Auto-resolved:  {auto_filled}")
    print(f"  Multiple hits:  {len(multi_match)} (fill scopus_id in verify_ID manually)")
    print(f"  Not found:      {not_found_ct}")

    if multi_match:
        print(f"\n  --- Multiple candidates — open verify_ID sheet and fill scopus_id manually ---")
        for name, univ, candidates in multi_match:
            print(f"\n  Faculty: {name} @ {univ}")
            for sid, dname, aff in candidates:
                print(f"    ID={sid:>15s}  name={dname:<30s}  affil={aff}")


# ---------------------------------------------------------------------------
# Normal pipeline mode
# ---------------------------------------------------------------------------
def run(indexes: list[str] | None = None):
    api_key = load_api_key()

    with open(CONFIG_PATH, encoding="utf-8") as f:
        universities = json.load(f)["universities"]

    if indexes:
        universities = [u for u in universities if u["index"] in indexes]

    for univ in universities:
        idx       = univ["index"]
        safe_name = univ["name"].replace(" ", "_").replace("/", "-")[:40]
        csv_path  = OUTPUT_DIR / f"{idx}_{safe_name}.csv"

        if not csv_path.exists():
            print(f"\n[{idx}] {univ['name']} — CSV not found, run scrape.py first")
            continue

        affiliation = univ.get("scopus_affil") or univ["name"]
        print(f"\n[{idx}] {univ['name']}")
        try:
            enrich(csv_path, affiliation=affiliation, api_key=api_key)
        except QuotaExhaustedError:
            print(f"\n  QUOTA EXHAUSTED — Scopus daily limit reached at [{idx}] {univ['name']}.")
            print(f"  Re-run with: python enrich.py --index {idx} (and all remaining indexes)")
            print(f"  Quota resets weekly (5,000 calls/week). Use a second API key or wait for the weekly reset.")
            break


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--index",  nargs="+", metavar="IDX",
                        help="Process only these university indexes (normal mode)")
    parser.add_argument("--verify", action="store_true",
                        help="Scan verify_ID sheet in output/2026_master.xlsx and resolve missing IDs")
    args = parser.parse_args()

    if args.verify:
        api_key = load_api_key()
        print(f"Verify mode — scanning {VERIFY_XLSX}")
        run_verify(VERIFY_XLSX, api_key)
    else:
        run(indexes=args.index)
